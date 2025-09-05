from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session, make_response, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import os
import uuid
from datetime import datetime, timedelta
import json
import csv
import io
import tempfile
from config import Config

app = Flask(__name__)
app.config.from_object(Config)

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Import models after db initialization
from models import User, Course, Room, TimeSlot, Exam, Constraint, Timetable

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.route('/')
def home():
    """Home page with attractive design and project overview"""
    return render_template('home.html')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        
        # Check if passwords match
        if password != confirm_password:
            flash('Passwords do not match!')
            return render_template('signup.html')
        
        # Check if user already exists
        if User.query.filter_by(email=email).first():
            flash('Email already registered!')
            return render_template('signup.html')
        
        if User.query.filter_by(username=username).first():
            flash('Username already taken!')
            return render_template('signup.html')
        
        # Create new user
        new_user = User(
            username=username,
            email=email,
            password_hash=generate_password_hash(password),
            role='user'  # Default role for new users
        )
        
        db.session.add(new_user)
        db.session.commit()
        
        flash('Account created successfully! Please login.')
        return redirect(url_for('login'))
    
    return render_template('signup.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        user = User.query.filter_by(email=email).first()
        
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid email or password')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('home'))

@app.route('/dashboard')
@login_required
def dashboard():
    """Main dashboard for timetabling operations"""
    courses = Course.query.all()
    rooms = Room.query.all()
    time_slots = TimeSlot.query.all()
    constraints = Constraint.query.all()
    
    return render_template('dashboard.html', 
                         courses=courses, 
                         rooms=rooms, 
                         time_slots=time_slots,
                         constraints=constraints)

@app.route('/courses', methods=['GET', 'POST'])
@login_required
def manage_courses():
    """Manage courses and their properties"""
    if request.method == 'POST':
        name = request.form.get('name')
        code = request.form.get('code')
        students = int(request.form.get('students', 0))
        duration = int(request.form.get('duration', 120))
        
        course = Course(name=name, code=code, students=students, duration=duration)
        db.session.add(course)
        db.session.commit()
        flash('Course added successfully!')
        return redirect(url_for('manage_courses'))
    
    courses = Course.query.all()
    return render_template('courses.html', courses=courses)

@app.route('/courses/export')
@login_required
def export_courses():
    """Export courses to CSV format"""
    try:
        courses = Course.query.all()
        
        # Create CSV content
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['Course Code', 'Course Name', 'Students', 'Duration (min)', 'Department'])
        
        # Write course data
        for course in courses:
            writer.writerow([
                course.code,
                course.name,
                course.students,
                course.duration,
                course.department
            ])
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=courses_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return response
        
    except Exception as e:
        flash(f'Export failed: {str(e)}', 'error')
        return redirect(url_for('manage_courses'))

@app.route('/rooms', methods=['GET', 'POST'])
@login_required
def manage_rooms():
    """Manage rooms and their capacities"""
    if request.method == 'POST':
        name = request.form.get('name')
        capacity = int(request.form.get('capacity', 0))
        building = request.form.get('building', '')
        
        room = Room(name=name, capacity=capacity, building=building)
        db.session.add(room)
        db.session.commit()
        flash('Room added successfully!')
        return redirect(url_for('manage_rooms'))
    
    rooms = Room.query.all()
    return render_template('rooms.html', rooms=rooms)

@app.route('/constraints', methods=['GET', 'POST'])
@login_required
def manage_constraints():
    """Manage scheduling constraints"""
    if request.method == 'POST':
        name = request.form.get('name')
        constraint_type = request.form.get('constraint_type')
        description = request.form.get('description')
        parameters = request.form.get('parameters', '{}')
        
        constraint = Constraint(
            name=name,
            constraint_type=constraint_type, 
            description=description, 
            parameters=parameters
        )
        db.session.add(constraint)
        db.session.commit()
        flash('Constraint added successfully!')
        return redirect(url_for('manage_constraints'))
    
    constraints = Constraint.query.all()
    return render_template('constraints.html', constraints=constraints)

@app.route('/optimize', methods=['POST'])
@login_required
def optimize_timetable():
    """Run the optimization algorithms"""
    try:
        # Get parameters from request
        algorithm = request.form.get('algorithm', 'hybrid')
        population_size = int(request.form.get('population_size', 50))
        generations = int(request.form.get('generations', 100))
        mutation_rate = float(request.form.get('mutation_rate', 0.1))
        temperature = float(request.form.get('temperature', 1000))
        cooling_rate = float(request.form.get('cooling_rate', 0.95))
        
        # Get data from database
        courses = Course.query.all()
        rooms = Room.query.all()
        time_slots = TimeSlot.query.all()
        constraints = Constraint.query.all()
        
        # Import algorithms here to avoid circular imports
        if algorithm == 'genetic':
            from algorithms.genetic_algorithm import GeneticAlgorithm
            optimizer = GeneticAlgorithm(courses, rooms, time_slots, constraints)
        elif algorithm == 'simulated_annealing':
            from algorithms.simulated_annealing import SimulatedAnnealing
            optimizer = SimulatedAnnealing(courses, rooms, time_slots, constraints)
        else:
            from algorithms.hybrid_optimizer import HybridOptimizer
            optimizer = HybridOptimizer(courses, rooms, time_slots, constraints)
        
        # Run optimization
        best_timetable, fitness_history = optimizer.optimize(
            population_size=population_size,
            generations=generations,
            mutation_rate=mutation_rate,
            temperature=temperature,
            cooling_rate=cooling_rate
        )
        
        # Save best timetable to database
        timetable = Timetable(
            data=json.dumps(best_timetable),
            fitness_score=best_timetable['fitness'],
            algorithm_used=algorithm,
            created_by=current_user.id,
            created_at=datetime.utcnow()
        )
        db.session.add(timetable)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'timetable': best_timetable,
            'fitness_history': fitness_history,
            'message': 'Optimization completed successfully!'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/timetables')
@login_required
def view_timetables():
    """View all generated timetables"""
    timetables = Timetable.query.order_by(Timetable.created_at.desc()).all()
    return render_template('timetables.html', timetables=timetables)

@app.route('/timetable/<int:timetable_id>')
@login_required
def view_timetable(timetable_id):
    """View a specific timetable"""
    timetable = Timetable.query.get_or_404(timetable_id)
    timetable_data = json.loads(timetable.data)
    
    # Parse the timetable entries for display
    entries = []
    if 'schedule' in timetable_data:
        for entry in timetable_data['schedule']:
            course = Course.query.get(entry.get('course_id'))
            room = Room.query.get(entry.get('room_id'))
            time_slot = TimeSlot.query.get(entry.get('time_slot_id'))
            
            if course and room and time_slot:
                entries.append({
                    'course': course,
                    'room': room,
                    'time_slot': time_slot
                })
    
    return render_template('timetable_detail.html', timetable=timetable, data=timetable_data, entries=entries)

@app.route('/export/<int:timetable_id>')
@login_required
def export_timetable(timetable_id):
    """Export timetable in various formats"""
    timetable = Timetable.query.get_or_404(timetable_id)
    format_type = request.args.get('format', 'csv')
    
    try:
        # Parse timetable data
        timetable_data = json.loads(timetable.data)
        
        # Debug logging
        print(f"Exporting timetable {timetable_id} in format {format_type}")
        print(f"Timetable data keys: {timetable_data.keys() if isinstance(timetable_data, dict) else 'Not a dict'}")
        
        if format_type.lower() == 'csv':
            return export_to_csv(timetable, timetable_data)
        elif format_type.lower() == 'excel':
            return export_to_excel(timetable, timetable_data)
        elif format_type.lower() == 'pdf':
            return export_to_pdf(timetable, timetable_data)
        else:
            return jsonify({'error': 'Unsupported format. Use csv, excel, or pdf'}), 400
            
    except Exception as e:
        print(f"Export error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Export failed: {str(e)}'}), 500

def export_to_csv(timetable, timetable_data):
    """Export timetable to CSV format"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(['Timetable Name', timetable.name])
    writer.writerow(['Algorithm Used', timetable.algorithm_used])
    writer.writerow(['Fitness Score', f"{timetable.fitness_score:.2f}"])
    writer.writerow(['Generated On', timetable.created_at.strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow([])  # Empty row
    
    # Write schedule header
    writer.writerow(['Day', 'Time Slot', 'Course Code', 'Course Name', 'Room', 'Building', 'Students', 'Duration (min)'])
    
    # Write schedule data - handle different data structures
    schedule_written = False
    
    # Try different possible data structures
    if isinstance(timetable_data, dict):
        if 'schedule' in timetable_data and isinstance(timetable_data['schedule'], list):
            for entry in timetable_data['schedule']:
                if isinstance(entry, dict):
                    course = Course.query.get(entry.get('course_id'))
                    room = Room.query.get(entry.get('room_id'))
                    time_slot = TimeSlot.query.get(entry.get('time_slot_id'))
                    
                    if course and room and time_slot:
                        writer.writerow([
                            time_slot.day,
                            f"{time_slot.start_time.strftime('%H:%M')} - {time_slot.end_time.strftime('%H:%M')}",
                            course.code,
                            course.name,
                            room.name,
                            room.building,
                            course.students,
                            course.duration
                        ])
                        schedule_written = True
        
        # If no schedule found, try to export available courses, rooms, and time slots
        if not schedule_written:
            courses = Course.query.all()
            rooms = Room.query.all()
            time_slots = TimeSlot.query.all()
            
            if courses and rooms and time_slots:
                writer.writerow(['Available Data Summary'])
                writer.writerow(['Courses:', len(courses)])
                writer.writerow(['Rooms:', len(rooms)])
                writer.writerow(['Time Slots:', len(time_slots)])
                writer.writerow([])
                
                writer.writerow(['Available Courses:'])
                writer.writerow(['Code', 'Name', 'Students', 'Duration'])
                for course in courses:
                    writer.writerow([course.code, course.name, course.students, course.duration])
    
    # If still no data, add a message
    if not schedule_written:
        writer.writerow(['No schedule data available for this timetable'])
    
    # Create response
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename="timetable_{timetable.id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    return response

def export_to_excel(timetable, timetable_data):
    """Export timetable to Excel format"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timetable"
        
        # Header styling
        header_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Add timetable information
        ws['A1'] = 'Timetable Information'
        ws['A1'].font = header_font
        ws['A2'] = 'Name:'
        ws['B2'] = timetable.name
        ws['A3'] = 'Algorithm:'
        ws['B3'] = timetable.algorithm_used
        ws['A4'] = 'Fitness Score:'
        ws['B4'] = f"{timetable.fitness_score:.2f}"
        ws['A5'] = 'Generated:'
        ws['B5'] = timetable.created_at.strftime('%Y-%m-%d %H:%M:%S')
        
        # Schedule header
        ws['A7'] = 'Examination Schedule'
        ws['A7'].font = header_font
        
        headers = ['Day', 'Time Slot', 'Course Code', 'Course Name', 'Room', 'Building', 'Students', 'Duration (min)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Add schedule data - handle different data structures
        row = 9
        schedule_written = False
        
        if isinstance(timetable_data, dict) and 'schedule' in timetable_data and isinstance(timetable_data['schedule'], list):
            for entry in timetable_data['schedule']:
                if isinstance(entry, dict):
                    course = Course.query.get(entry.get('course_id'))
                    room = Room.query.get(entry.get('room_id'))
                    time_slot = TimeSlot.query.get(entry.get('time_slot_id'))
                    
                    if course and room and time_slot:
                        ws.cell(row=row, column=1, value=time_slot.day)
                        ws.cell(row=row, column=2, value=f"{time_slot.start_time.strftime('%H:%M')} - {time_slot.end_time.strftime('%H:%M')}")
                        ws.cell(row=row, column=3, value=course.code)
                        ws.cell(row=row, column=4, value=course.name)
                        ws.cell(row=row, column=5, value=room.name)
                        ws.cell(row=row, column=6, value=room.building)
                        ws.cell(row=row, column=7, value=course.students)
                        ws.cell(row=row, column=8, value=course.duration)
                        row += 1
                        schedule_written = True
        
        # If no schedule data, add available courses
        if not schedule_written:
            courses = Course.query.all()
            if courses:
                ws.cell(row=row, column=1, value="Available Courses:")
                row += 1
                for course in courses:
                    ws.cell(row=row, column=1, value="N/A")
                    ws.cell(row=row, column=2, value="N/A")
                    ws.cell(row=row, column=3, value=course.code)
                    ws.cell(row=row, column=4, value=course.name)
                    ws.cell(row=row, column=5, value="N/A")
                    ws.cell(row=row, column=6, value="N/A")
                    ws.cell(row=row, column=7, value=course.students)
                    ws.cell(row=row, column=8, value=course.duration)
                    row += 1
            else:
                ws.cell(row=row, column=1, value="No schedule data available")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f"timetable_{timetable.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except ImportError:
        return jsonify({'error': 'Excel export requires openpyxl package. Please install it: pip install openpyxl'}), 500

def export_to_pdf(timetable, timetable_data):
    """Export timetable to PDF format"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        
        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        
        # Create PDF document
        doc = SimpleDocTemplate(temp_file.name, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        story.append(Paragraph("Examination Timetable", title_style))
        
        # Timetable information
        info_data = [
            ['Timetable Name:', timetable.name],
            ['Algorithm Used:', timetable.algorithm_used],
            ['Fitness Score:', f"{timetable.fitness_score:.2f}"],
            ['Generated On:', timetable.created_at.strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(info_table)
        story.append(Spacer(1, 20))
        
        # Schedule table
        schedule_data = [['Day', 'Time', 'Course', 'Room', 'Building', 'Students', 'Duration']]
        
        schedule_written = False
        if isinstance(timetable_data, dict) and 'schedule' in timetable_data and isinstance(timetable_data['schedule'], list):
            for entry in timetable_data['schedule']:
                if isinstance(entry, dict):
                    course = Course.query.get(entry.get('course_id'))
                    room = Room.query.get(entry.get('room_id'))
                    time_slot = TimeSlot.query.get(entry.get('time_slot_id'))
                    
                    if course and room and time_slot:
                        schedule_data.append([
                            time_slot.day,
                            f"{time_slot.start_time.strftime('%H:%M')}-{time_slot.end_time.strftime('%H:%M')}",
                            f"{course.code}\n{course.name}",
                            room.name,
                            room.building,
                            str(course.students),
                            f"{course.duration}m"
                        ])
                        schedule_written = True
        
        # If no schedule data, add available courses
        if not schedule_written:
            courses = Course.query.all()
            if courses:
                for course in courses:
                    schedule_data.append([
                        "N/A",
                        "N/A", 
                        f"{course.code}\n{course.name}",
                        "N/A",
                        "N/A",
                        str(course.students),
                        f"{course.duration}m"
                    ])
            else:
                schedule_data.append(["No schedule data available", "", "", "", "", "", ""])
        
        schedule_table = Table(schedule_data, colWidths=[0.8*inch, 1*inch, 1.8*inch, 0.8*inch, 1*inch, 0.6*inch, 0.6*inch])
        schedule_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(Paragraph("Examination Schedule", styles['Heading2']))
        story.append(Spacer(1, 12))
        story.append(schedule_table)
        
        # Build PDF
        doc.build(story)
        temp_file.close()
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f"timetable_{timetable.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
            mimetype='application/pdf'
        )
        
    except ImportError as e:
        print(f"PDF Export Error - ImportError: {e}")
        return jsonify({'error': 'PDF export requires reportlab package. Please install it: pip install reportlab'}), 500
    except Exception as e:
        print(f"PDF Export Error - General Exception: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'PDF export failed: {str(e)}'}), 500

@app.route('/test-export')
@login_required
def test_export():
    """Test export functionality with sample data"""
    # Create a test timetable with sample data
    sample_data = {
        'schedule': [],
        'fitness': 85.5,
        'algorithm': 'genetic',
        'test': True
    }
    
    # Get some sample courses for testing
    courses = Course.query.limit(3).all()
    rooms = Room.query.limit(3).all() 
    time_slots = TimeSlot.query.limit(3).all()
    
    if courses and rooms and time_slots:
        for i in range(min(len(courses), len(rooms), len(time_slots))):
            sample_data['schedule'].append({
                'course_id': courses[i].id,
                'room_id': rooms[i].id,
                'time_slot_id': time_slots[i].id
            })
    
    # Create a temporary timetable object for testing
    class TestTimetable:
        def __init__(self):
            self.id = 999
            self.name = "Test Export Timetable"
            self.algorithm_used = "genetic"
            self.fitness_score = 85.5
            self.created_at = datetime.now()
    
    test_timetable = TestTimetable()
    format_type = request.args.get('format', 'csv')
    
    try:
        if format_type.lower() == 'csv':
            return export_to_csv(test_timetable, sample_data)
        elif format_type.lower() == 'excel':
            return export_to_excel(test_timetable, sample_data)
        elif format_type.lower() == 'pdf':
            return export_to_pdf(test_timetable, sample_data)
        else:
            return jsonify({'error': 'Unsupported format. Use csv, excel, or pdf'}), 400
    except Exception as e:
        return jsonify({'error': f'Test export failed: {str(e)}'}), 500

@app.route('/export/<int:timetable_id>')
@login_required
def export_timetable(timetable_id):
    timetable = Timetable.query.get_or_404(timetable_id)
    format_type = request.args.get('format', 'csv')
    
    print(f"Export request: timetable_id={timetable_id}, format={format_type}")
    
    try:
        timetable_data = json.loads(timetable.data) if timetable.data else {}
        print(f"Timetable data loaded: {type(timetable_data)}")
        
        if format_type.lower() == 'csv':
            return export_to_csv(timetable, timetable_data)
        elif format_type.lower() == 'excel':
            return export_to_excel(timetable, timetable_data)
        elif format_type.lower() == 'pdf':
            print("Attempting PDF export...")
            return export_to_pdf(timetable, timetable_data)
        else:
            return jsonify({'error': 'Unsupported format. Use csv, excel, or pdf'}), 400
            
    except Exception as e:
        print(f"Export route error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Export failed: {str(e)}'}), 500

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, host='0.0.0.0', port=5000)
